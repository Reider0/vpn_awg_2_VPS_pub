import asyncpg
import os
import asyncio
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from utils import dt_to_moscow

CONFIGS_DIR = Path("/volumes/configs")
WG_CONF_PATH = Path("/volumes/wireguard/wg0.conf")

class Database:
    def __init__(self):
        self.pool = None
        self.database_url = os.getenv("DATABASE_URL", "postgres://vpn:vpnpass@postgres:5432/vpndb")

    async def connect(self):
        for i in range(5):
            try:
                self.pool = await asyncpg.create_pool(dsn=self.database_url)
                await self.init_tables()
                await self._check_migrations()
                print("Успешное подключение к БД")
                break
            except Exception as e:
                print(f"Попытка подключения к БД {i+1} неудачна: {e}")
                await asyncio.sleep(2)

    async def init_tables(self):
        pass

    async def _check_migrations(self):
        try:
            await self.execute("""
                CREATE TABLE IF NOT EXISTS user_tg_links (
                    uuid TEXT REFERENCES users(uuid) ON DELETE CASCADE,
                    tg_id BIGINT,
                    UNIQUE(uuid, tg_id)
                );
            """)
            await self.execute("""
                CREATE TABLE IF NOT EXISTS events_log (
                    id SERIAL PRIMARY KEY,
                    timestamp TIMESTAMP DEFAULT NOW(),
                    event_type TEXT,
                    message TEXT
                );
            """)
            await self.execute("""
                CREATE TABLE IF NOT EXISTS support_tickets (
                    id SERIAL PRIMARY KEY,
                    user_uuid TEXT REFERENCES users(uuid) ON DELETE CASCADE,
                    message TEXT,
                    status TEXT DEFAULT 'open',
                    created_at TIMESTAMP DEFAULT NOW()
                );
            """)
            await self.execute("""
                CREATE TABLE IF NOT EXISTS user_ips (
                    uuid TEXT REFERENCES users(uuid) ON DELETE CASCADE,
                    ip TEXT,
                    status TEXT DEFAULT 'pending',
                    first_seen TIMESTAMP DEFAULT NOW(),
                    last_seen TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (uuid, ip)
                );
            """)

            res_tg = await self.fetch_all("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='tg_id';")
            if res_tg:
                await self.execute("INSERT INTO user_tg_links (uuid, tg_id) SELECT uuid, tg_id FROM users WHERE tg_id IS NOT NULL ON CONFLICT DO NOTHING;")

            res_exp = await self.fetch_all("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='expires_at';")
            if not res_exp:
                await self.execute("ALTER TABLE users ADD COLUMN expires_at TIMESTAMP;")
                
            res_act = await self.fetch_all("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='is_active';")
            if not res_act:
                await self.execute("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT TRUE;")

            res_last = await self.fetch_all("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='last_active_at';")
            if not res_last:
                await self.execute("ALTER TABLE users ADD COLUMN last_active_at TIMESTAMP;")

        except Exception as e:
            print(f"Migration error: {e}")

    async def track_user_ip(self, uuid, ip):
        existing = await self.fetch_val("SELECT status FROM user_ips WHERE uuid=$1 AND ip=$2", uuid, ip)
        if not existing:
            await self.execute("INSERT INTO user_ips (uuid, ip) VALUES ($1, $2)", uuid, ip)
            return True 
        else:
            await self.execute("UPDATE user_ips SET last_seen = NOW() WHERE uuid=$1 AND ip=$2", uuid, ip)
            if existing == 'pending':
                await self.execute("UPDATE user_ips SET status = 'trusted' WHERE uuid=$1 AND status = 'pending' AND first_seen < NOW() - INTERVAL '48 HOURS'", uuid)
            return False 

    async def get_user_ips(self, uuid):
        query = "SELECT ip, status, first_seen, last_seen FROM user_ips WHERE uuid=$1 ORDER BY last_seen DESC LIMIT 8"
        return await self.fetch_all(query, uuid)

    async def log_event(self, event_type, message):
        try:
            query = "INSERT INTO events_log (event_type, message) VALUES ($1, $2)"
            await self.execute(query, event_type, message)
        except Exception as e:
            print(f"Failed to log event (ignored): {e}")

    async def execute(self, query, *args):
        for i in range(3):
            try:
                if not self.pool: await self.connect()
                async with self.pool.acquire() as conn:
                    return await conn.execute(query, *args)
            except Exception as e:
                print(f"DB Execute Error: {e}. Retrying {i+1}/3...")
                await asyncio.sleep(2)
                if i == 2: raise

    async def fetch_all(self, query, *args):
        for i in range(3):
            try:
                if not self.pool: await self.connect()
                async with self.pool.acquire() as conn:
                    return await conn.fetch(query, *args)
            except Exception as e:
                print(f"DB Fetch Error: {e}. Retrying {i+1}/3...")
                await asyncio.sleep(2)
                if i == 2: raise
    
    async def fetch_val(self, query, *args):
        for i in range(3):
            try:
                if not self.pool: await self.connect()
                async with self.pool.acquire() as conn:
                    return await conn.fetchval(query, *args)
            except Exception as e:
                print(f"DB Fetchval Error: {e}. Retrying {i+1}/3...")
                await asyncio.sleep(2)
                if i == 2: raise

    async def device_set(self, uuid):
        query = "SELECT device FROM users WHERE uuid=$1"
        result = await self.fetch_all(query, uuid)
        if result and result[0]["device"]:
            return True
        return False

    async def get_all_users(self):
        query = """
            SELECT u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at,
                   ARRAY_REMOVE(ARRAY_AGG(l.tg_id), NULL) as tg_ids
            FROM users u
            LEFT JOIN user_tg_links l ON u.uuid = l.uuid
            GROUP BY u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at
            ORDER BY u.created_at DESC
        """
        return await self.fetch_all(query)
    
    async def get_user_by_uuid(self, uuid):
        query = """
            SELECT u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at,
                   ARRAY_REMOVE(ARRAY_AGG(l.tg_id), NULL) as tg_ids
            FROM users u
            LEFT JOIN user_tg_links l ON u.uuid = l.uuid
            WHERE u.uuid=$1
            GROUP BY u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at
        """
        rows = await self.fetch_all(query, uuid)
        return rows[0] if rows else None

    async def get_users_by_tg_id(self, tg_id):
        query = """
            SELECT u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at,
                   ARRAY_REMOVE(ARRAY_AGG(l.tg_id), NULL) as tg_ids 
            FROM users u
            JOIN user_tg_links l ON u.uuid = l.uuid
            WHERE l.tg_id=$1
            GROUP BY u.id, u.name, u.uuid, u.device, u.is_active, u.expires_at, u.created_at, u.first_connected_at, u.last_active_at
        """
        return await self.fetch_all(query, tg_id)

    async def link_user_telegram(self, uuid, tg_id):
        await self.execute("INSERT INTO user_tg_links (uuid, tg_id) VALUES ($1, $2) ON CONFLICT DO NOTHING", uuid, tg_id)
        await self.log_event("Link TG", f"Linked TG {tg_id} to key {uuid}")

    async def unlink_user_telegram(self, uuid, tg_id):
        await self.execute("DELETE FROM user_tg_links WHERE uuid=$1 AND tg_id=$2", uuid, tg_id)
        await self.log_event("Unlink TG", f"Unlinked TG {tg_id} from key {uuid}")

    async def get_all_tg_ids(self):
        query = "SELECT DISTINCT tg_id FROM user_tg_links"
        rows = await self.fetch_all(query)
        return [r['tg_id'] for r in rows]

    async def save_stats(self, uuid, rx, tx):
        user = await self.get_user_by_uuid(uuid)
        if user:
            query = "INSERT INTO stats (user_uuid, bytes_in, bytes_out, last_seen) VALUES ($1, $2, $3, NOW())"
            await self.execute(query, uuid, rx, tx)

    async def get_stats_24h(self):
        query = """
            SELECT s.user_uuid, u.name, s.bytes_in, s.bytes_out, s.last_seen 
            FROM stats s
            JOIN users u ON s.user_uuid = u.uuid
            WHERE s.last_seen >= NOW() - INTERVAL '24 HOURS'
            ORDER BY s.last_seen ASC
        """
        return await self.fetch_all(query)

    async def cleanup_old_logs(self, days=7):
        await self.execute(f"DELETE FROM events_log WHERE timestamp < NOW() - INTERVAL '{days} DAYS'")
        await self.execute(f"DELETE FROM stats WHERE last_seen < NOW() - INTERVAL '{days} DAYS'")
        await self.execute(f"DELETE FROM user_ips WHERE status = 'pending' AND last_seen < NOW() - INTERVAL '7 DAYS'")

    async def set_setting(self, key, value):
        await self.execute("INSERT INTO settings (key, value) VALUES ($1, $2) ON CONFLICT (key) DO UPDATE SET value = $2", key, str(value))

    async def get_setting(self, key):
        return await self.fetch_val("SELECT value FROM settings WHERE key=$1", key)

    async def export_to_excel(self, path):
        users = await self.get_all_users()
        wb = Workbook()
        
        ws = wb.active
        ws.title = "VPN Users"

        headers =["Имя", "UUID", "Устройство", "Статус", "Годен до", "TG IDs", "Создано", "Первый Вход", "Посл. Активность", "Конфиг (Текст)", "QR Код"]
        ws.append(headers)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 45
        ws.column_dimensions['K'].width = 25

        for index, u in enumerate(users, start=2):
            status_text = "Активен" if u['is_active'] else "Пауза"
            exp_text = dt_to_moscow(u['expires_at']).strftime("%Y-%m-%d %H:%M") if u['expires_at'] else "Навсегда"
            
            ws.cell(row=index, column=1, value=u['name'])
            ws.cell(row=index, column=2, value=u['uuid'])
            ws.cell(row=index, column=3, value=u['device'] or "")
            ws.cell(row=index, column=4, value=status_text)
            ws.cell(row=index, column=5, value=exp_text)
            
            tg_ids_str = ", ".join(map(str, u.get('tg_ids',[])))
            ws.cell(row=index, column=6, value=tg_ids_str)
            
            ws.cell(row=index, column=7, value=dt_to_moscow(u['created_at']).strftime("%Y-%m-%d %H:%M") if u['created_at'] else "")
            ws.cell(row=index, column=8, value=dt_to_moscow(u['first_connected_at']).strftime("%Y-%m-%d %H:%M") if u['first_connected_at'] else "")
            ws.cell(row=index, column=9, value=dt_to_moscow(u['last_active_at']).strftime("%Y-%m-%d %H:%M") if u.get('last_active_at') else "Нет данных")

            conf_path = CONFIGS_DIR / f"{u['name']}.conf"
            if not conf_path.exists():
                conf_path = CONFIGS_DIR / f"{u['name']}_Full.conf"
                
            if conf_path.exists():
                with open(conf_path, "r") as f:
                    conf_text = f.read()
                cell = ws.cell(row=index, column=10, value=conf_text)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            qr_path = CONFIGS_DIR / f"{u['name']}.png"
            if not qr_path.exists():
                qr_path = CONFIGS_DIR / f"{u['name']}_Full.png"

            if qr_path.exists():
                try:
                    img = ExcelImage(str(qr_path))
                    img.width = 150
                    img.height = 150
                    ws.add_image(img, f"K{index}")
                    ws.row_dimensions[index].height = 120
                except Exception as e:
                    ws.cell(row=index, column=11, value=f"Ошибка картинки: {e}")
            else:
                ws.row_dimensions[index].height = 120

        ws_events = wb.create_sheet(title="System Events (Last 7 Days)")
        ws_events.append(["Дата и Время (МСК)", "Тип события", "Сообщение"])
        ws_events.column_dimensions['A'].width = 25
        ws_events.column_dimensions['B'].width = 20
        ws_events.column_dimensions['C'].width = 70
        
        logs = await self.fetch_all("SELECT timestamp, event_type, message FROM events_log ORDER BY timestamp DESC")
        for row in logs:
            ws_events.append([dt_to_moscow(row['timestamp']).strftime("%Y-%m-%d %H:%M:%S"), row['event_type'], row['message']])

        ws_wg = wb.create_sheet(title="wg0.conf")
        ws_wg.column_dimensions['A'].width = 120
        if WG_CONF_PATH.exists():
            with open(WG_CONF_PATH, "r") as f:
                wg_text = f.read()
            cell = ws_wg.cell(row=1, column=1, value=wg_text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws_wg.cell(row=1, column=1, value="Файл wg0.conf не найден в контейнере!")

        wb.save(path)
        return path

    async def export_logs_to_excel(self, path):
        users = await self.get_all_users()
        wb = Workbook()
        wb.remove(wb.active)

        if not users:
            ws = wb.create_sheet("Empty")
            ws.append(["Нет пользователей"])
            wb.save(path)
            return path

        for u in users:
            safe_name = "".join([c for c in u['name'] if c.isalnum() or c == '_'])[:30]
            if not safe_name: safe_name = u['uuid'][:8]
            
            ws = wb.create_sheet(title=safe_name)
            ws.append(["Время МСК", "Скачано (MB)", "Отправлено (MB)", "Активность (MB)"])
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30

            query = "SELECT bytes_in, bytes_out, last_seen FROM stats WHERE user_uuid=$1 ORDER BY last_seen ASC"
            user_stats = await self.fetch_all(query, u['uuid'])
            
            prev_total = 0
            for row in user_stats:
                total_bytes = row['bytes_in'] + row['bytes_out']
                delta_bytes = total_bytes - prev_total
                if delta_bytes < 0: delta_bytes = total_bytes
                if prev_total == 0: delta_bytes = 0
                prev_total = total_bytes
                
                ws.append([
                    dt_to_moscow(row['last_seen']).strftime("%Y-%m-%d %H:%M:%S"),
                    round(row['bytes_in'] / (1024 * 1024), 2),
                    round(row['bytes_out'] / (1024 * 1024), 2),
                    round(delta_bytes / (1024 * 1024), 2)
                ])
                
        wb.save(path)
        return path

db = Database()